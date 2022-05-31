from tkinter import *
import openpyxl as xl
from tkinter import messagebox
import admin_form as admin

def logging_in():
  root = Tk()
  root.title("Food Frenzy")
  root.geometry("550x400+100+50")
  root.resizable(False, False)
  positionRight = int(root.winfo_screenwidth() / 2 - 300 / 2)
  positionDown = int(root.winfo_screenheight() / 2.2 - 300 / 2)
  root.geometry("+{}+{}".format(positionRight, positionDown))
  
  global count
  count =1
  
  wb = xl.load_workbook('accounts.xlsx')
  sheet = wb["Sheet1"]

  user = []
  pas = []
  for row in range(1, sheet.max_row + 1):
      cell_user = sheet.cell(row, 1)
      cell_pass = sheet.cell(row, 2)
      user.append(cell_user.value)
      pas.append(cell_pass.value)
  
  def login_function():
      global count
      username = username_ent.get()
      password = password_ent.get()

      if username == 'admin' and password == "12345":
          root.destroy()
          admin.administration()
          
      else:
          if password in pas and username in user:
              user_data = user.index(username)
              pass_data = pas.index(password)
              if user_data == pass_data:
                  root.destroy()
                  import game
              else:
                  if count < 3:
                      count += 1
                      messagebox.showwarning("", "Username and Password did not matched!")
                  else:
                      count = 1
                      messagebox.showerror("", "You consecutively failed to log-in. Please try again later.")
                      root.destroy()
          elif username not in user or password not in pas:
              if count < 3:
                  count += 1
                  messagebox.showwarning("", "Username and Password did not matched!")
              else:
                  count = 1
                  messagebox.showerror("", "You consecutively failed to log-in. Please try again later.")
                  root.destroy()
                
  #Elements
  sign_in = Label(root, text="Sign In")
  username_lbl = Label(root, text="Username")
  password_lbl = Label(root, text="Password")
  username_ent = Entry(root)
  password_ent = Entry(root)
  log_in_btn = Button(root, text="Log In", command=lambda: log_in())







  #Display
  sign_in.place(x=150, y=5)
  username_lbl.place(x=75, y=10)
  password_lbl.place(x=75,y=15)
  username_ent.place(x=85, y=10)
  password_ent.place(x=85,y=15)
  log_in_btn.place(x=150, y=250)
  register.place(x=250, y=280)
  
  root.mainloop()

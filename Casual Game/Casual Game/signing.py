from tkinter import *
import openpyxl as xl
from tkinter import messagebox

def signing_up():
  root = Tk()
  root.title("Food Frenzy")
  root.geometry("550x400+100+50")
  root.resizable(False, False)
  positionRight = int(root.winfo_screenwidth() / 2 - 300 / 2)
  positionDown = int(root.winfo_screenheight() / 2.2 - 300 / 2)
  root.geometry("+{}+{}".format(positionRight, positionDown))

  # Database
  wb = xl.load_workbook("accounts.xlsx")
  sheet = wb["Sheet1"]


  def signed_up():

      username = username_ent.get()
      password = password_ent.get()

      cell_user = sheet.cell(sheet.max_row + 1, 1)
      cell_pass = sheet.cell(sheet.max_row, 2)

      cell_user.value = username
      cell_pass.value = password
      if password.isdecimal() or (username == "" and password == ""):
          messagebox.showerror("","Password must contain at least 1 character")
      else:
          wb.save("accounts.xlsx")
          messagebox.showinfo("Signed Up", "Thanks for signing up!")
          root.destroy()



  username_lbl = Label(root, text="Username :", font=("Arial Narrow",15,"bold"))
  password_lbl = Label(root, text="Password :", font=("Arial Narrow",15,"bold"))
  username_ent = Entry(root)
  password_ent = Entry(root)
  reg_but = Button(root, text="Sign Up", font=("Arial Narrow",15,"bold"), command=lambda: signed_up())



  username_lbl.place(x=150, y=165)
  password_lbl.place(x=150, y=200)
  username_ent.place(x=250, y=165, width=147, height=25)
  password_ent.place(x=250, y=200, width=147, height=25)
  reg_but.place(x=238, y=250)

  root.mainloop()
